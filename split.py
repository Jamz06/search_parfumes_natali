from openpyxl import load_workbook, Workbook

def split_excel_by_rows(input_file, rows_per_file=100):
    wb = load_workbook(input_file)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    total_rows = len(rows)
    part = 1

    for start_row in range(0, total_rows, rows_per_file):
        new_wb = Workbook()
        new_ws = new_wb.active

        chunk = rows[start_row:start_row+rows_per_file]
        for row in chunk:
            new_ws.append(row)

        new_wb.save(f'part_{part}.xlsx')
        part += 1

# Usage
split_excel_by_rows('your_file.xlsx', 100)