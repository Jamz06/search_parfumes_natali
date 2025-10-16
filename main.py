import os
import time
import datetime

import openpyxl

from duckduckgo_search import DDGS
from duckduckgo_search.exceptions import RatelimitException

SLEEP = 5
# С какой строки начать
FIRST_ROW = 8
# Колонка с текстом для поиска
SEARCH_COLUMN = 3
# Колонка с URL
URL_COLUMN = 11
# Сколько максимально нужно картинок
MAX_RESULTS = 2
# Сохранять после обработки N строк:
SAVE_AFTER_N_ROWS = 100

def init():
    """
    Инициализация
    """
    # Проверить наличие дирректории для результатов
    if not os.path.exists('results'):
        os.makedirs('results')
    
    if not os.path.exists('xls'):
        os.makedirs('xls')
    
def list_files():
    return os.listdir('xls')
        
def search_for_image(keyword):
    with DDGS() as ddgs:
        results = ddgs.images(
            keyword,
            max_results=MAX_RESULTS,
            region="ru-ru"        
        )
        image_urls = [r['image'] for r in results]

    return(image_urls)

def main():
    
    init()
    input_files = list_files()
    if len(input_files) == 0:
        print('Нет входных файлов')
        print('Сохрани таблицу в каталоге xls')
        return

    # Возьму один файл, лучше не обходить список, т.к. файлы разные
    input_file = input_files[0]
    if not input_file.endswith('.xlsx'):
        print('Файл должен быть в формате xslx')
        return

    timestamp = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    result_filename = f'results/results_{timestamp}.xlsx'
    
    workbook = openpyxl.load_workbook(f'xls/{input_file}')
    sheet = workbook.active
    
    print(f'Всего строк: {sheet.max_row}')
    seconds = sheet.max_row * (SLEEP+0.5)
    print(f'Примерно займет времени: {seconds // 60} минут {seconds % 60} секунд')

        
    for row in range(FIRST_ROW, sheet.max_row + 1):
        if row % SAVE_AFTER_N_ROWS == 0:
            workbook.save(result_filename)
            print(f'Промежуточный Результат сохранен в {result_filename}')
        
        keyword = sheet.cell(row=row, column=SEARCH_COLUMN).value
        
        print(f'Строка {row} поиск: {keyword}')
        
        search_error = True
        
        try:
            while search_error == True:
                try:
                    image_urls = search_for_image(keyword)
                except RatelimitException as err:        
                    print(err)
                    print("Ошибка предела числа запросов поиска, жду 15 секунд")
                    time.sleep(15)
                    continue
                
                print(f'Найдено картинок: {image_urls}')
                search_error = False
                
            
        except Exception as err:
            print("Произошла ошибка:")
            print(err)
            print("Сохраняю результат")
            result_filename = f'results/results_{timestamp}_error_{row}.xlsx'
            break
        
        
        
        # Сохранить урлы картинок
        for url_idx in range(len(image_urls)):
            sheet.cell(row=row, column=URL_COLUMN+url_idx).value = image_urls[url_idx]
            
        time.sleep(SLEEP)

    workbook.save(result_filename)
    print(f'Результат сохранен в {result_filename}')
    
if __name__ == '__main__':
    main()